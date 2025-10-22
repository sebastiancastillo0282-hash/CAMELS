"""Utilities to produce consolidated CAMELS exports."""
from __future__ import annotations

import csv
import json
import sqlite3
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Mapping, Sequence, Tuple

from openpyxl import Workbook


@dataclass(slots=True)
class ExportSummary:
    """Details about the generated export files."""

    portfolio_rows: int
    indicator_rows: int
    files: List[Path]


class ExportGenerator:
    """Create consolidated CSV/Excel artifacts from SQLite outputs."""

    def __init__(self, sqlite_path: Path, output_dir: Path) -> None:
        self.sqlite_path = sqlite_path
        self.output_dir = output_dir

    def generate(self, run_id: str) -> ExportSummary:
        """Generate CSV/Excel exports for *run_id* results."""

        portfolio = self._portfolio_rows(run_id)
        indicators = self._indicator_rows(run_id)

        if not portfolio and not indicators:
            return ExportSummary(0, 0, [])

        self.output_dir.mkdir(parents=True, exist_ok=True)

        portfolio_csv = self.output_dir / f"camels_portfolio_{run_id}.csv"
        indicator_csv = self.output_dir / f"camels_indicators_{run_id}.csv"
        workbook_path = self.output_dir / f"camels_report_{run_id}.xlsx"

        self._write_csv(portfolio_csv, portfolio)
        self._write_csv(indicator_csv, indicators)
        self._write_workbook(workbook_path, portfolio, indicators)

        return ExportSummary(
            portfolio_rows=len(portfolio),
            indicator_rows=len(indicators),
            files=[portfolio_csv, indicator_csv, workbook_path],
        )

    # ------------------------------------------------------------------
    # Data retrieval helpers
    # ------------------------------------------------------------------

    def _portfolio_rows(self, run_id: str) -> List[Dict[str, object]]:
        query = (
            """
            SELECT s.bank_id,
                   b.name AS bank_name,
                   b.country,
                   b.regulator,
                   s.score,
                   s.rating,
                   s.period,
                   s.details
              FROM scores s
              JOIN banks b ON b.bank_id = s.bank_id
             WHERE s.run_id = ?
             ORDER BY s.score DESC
            """
        )
        with sqlite3.connect(self.sqlite_path) as connection:
            connection.row_factory = sqlite3.Row
            rows = connection.execute(query, (run_id,)).fetchall()

        results: List[Dict[str, object]] = []
        for row in rows:
            metadata = self._safe_json(row["details"])
            results.append(
                {
                    "bank_id": row["bank_id"],
                    "bank_name": row["bank_name"],
                    "country": row["country"],
                    "regulator": row["regulator"],
                    "score": row["score"],
                    "rating": row["rating"],
                    "period": row["period"],
                    "metadata": metadata,
                }
            )
        return results

    def _indicator_rows(self, run_id: str) -> List[Dict[str, object]]:
        query = (
            """
            SELECT i.bank_id,
                   banks.name AS bank_name,
                   banks.country,
                   banks.regulator,
                   i.indicator_id,
                   ind.name AS indicator_name,
                   i.pillar,
                   i.value,
                   i.score,
                   i.rating,
                   i.weight,
                   i.period,
                   i.unit,
                   i.source_id,
                   i.normalization_run_id,
                   i.details
              FROM indicator_scores i
              JOIN banks ON banks.bank_id = i.bank_id
         LEFT JOIN indicators ind ON ind.indicator_id = i.indicator_id
             WHERE i.run_id = ?
             ORDER BY banks.name, i.pillar, ind.name
            """
        )
        with sqlite3.connect(self.sqlite_path) as connection:
            connection.row_factory = sqlite3.Row
            rows = connection.execute(query, (run_id,)).fetchall()

        indicators: List[Dict[str, object]] = []
        ingestion_keys: List[Tuple[str, str | None]] = []
        for row in rows:
            details = self._safe_json(row["details"])
            source_meta = details.get("source_metadata") or {}
            ingestion_run = source_meta.get("source_run")
            ingestion_keys.append((row["source_id"], ingestion_run))
            indicators.append(
                {
                    "bank_id": row["bank_id"],
                    "bank_name": row["bank_name"],
                    "country": row["country"],
                    "regulator": row["regulator"],
                    "indicator_id": row["indicator_id"],
                    "indicator_name": row["indicator_name"],
                    "pillar": row["pillar"],
                    "value": row["value"],
                    "score": row["score"],
                    "rating": row["rating"],
                    "weight": row["weight"],
                    "period": row["period"],
                    "unit": row["unit"],
                    "source_id": row["source_id"],
                    "normalization_run_id": row["normalization_run_id"],
                    "metadata": details,
                    "ingestion_run_id": ingestion_run,
                }
            )

        ingestion_lookup = self._ingestion_lookup(ingestion_keys)
        for entry in indicators:
            source_id = entry.get("source_id")
            ingestion_run = entry.get("ingestion_run_id")
            info = None
            if source_id:
                info = ingestion_lookup.get((source_id, ingestion_run)) or ingestion_lookup.get((source_id, None))
            if info:
                entry["source_url"] = info.get("url")
                entry["document_path"] = info.get("local_path")
                entry["checksum"] = info.get("checksum")
            else:
                entry["source_url"] = None
                entry["document_path"] = None
                entry["checksum"] = (
                    entry["metadata"].get("source_metadata", {}).get("checksum")
                    if isinstance(entry["metadata"], Mapping)
                    else None
                )
        return indicators

    # ------------------------------------------------------------------
    # Output helpers
    # ------------------------------------------------------------------

    def _write_csv(self, path: Path, rows: Sequence[Mapping[str, object]]) -> None:
        fieldnames = self._determine_fieldnames(rows)
        with path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=fieldnames)
            writer.writeheader()
            for row in rows:
                serialised = {
                    key: json.dumps(value, ensure_ascii=False)
                    if isinstance(value, (dict, list))
                    else value
                    for key, value in row.items()
                }
                writer.writerow(serialised)

    def _write_workbook(
        self,
        path: Path,
        portfolio: Sequence[Mapping[str, object]],
        indicators: Sequence[Mapping[str, object]],
    ) -> None:
        workbook = Workbook()
        scores_sheet = workbook.active
        scores_sheet.title = "Scores"
        self._write_sheet(scores_sheet, portfolio)
        indicator_sheet = workbook.create_sheet("Indicators")
        self._write_sheet(indicator_sheet, indicators)
        workbook.save(path)

    def _write_sheet(self, worksheet, rows: Sequence[Mapping[str, object]]) -> None:
        from openpyxl.utils import get_column_letter

        fieldnames = self._determine_fieldnames(rows)
        if not fieldnames:
            fieldnames = ["message"]
            rows = [{"message": "No data available"}]
        worksheet.append(fieldnames)
        for row in rows:
            worksheet.append(
                [
                    json.dumps(value, ensure_ascii=False)
                    if isinstance(value, (dict, list))
                    else value
                    for value in (row.get(field) for field in fieldnames)
                ]
            )
        for index, _ in enumerate(fieldnames, start=1):
            column = get_column_letter(index)
            worksheet.column_dimensions[column].width = 18

    # ------------------------------------------------------------------
    # Utility helpers
    # ------------------------------------------------------------------

    def _safe_json(self, payload: object) -> Dict[str, object]:
        if not payload:
            return {}
        if isinstance(payload, dict):
            return payload
        try:
            return json.loads(payload)
        except (TypeError, json.JSONDecodeError):
            return {"raw": payload}

    def _determine_fieldnames(
        self, rows: Sequence[Mapping[str, object]]
    ) -> List[str]:
        fieldnames: List[str] = []
        for row in rows:
            for key in row.keys():
                if key not in fieldnames:
                    fieldnames.append(key)
        return fieldnames

    def _ingestion_lookup(
        self, keys: Iterable[Tuple[str | None, str | None]]
    ) -> Dict[Tuple[str, str | None], Dict[str, object]]:
        valid_sources = {key[0] for key in keys if key[0]}
        if not valid_sources:
            return {}
        placeholders = ",".join("?" for _ in valid_sources)
        query = (
            f"""
            SELECT run_id, source_id, url, local_path, checksum, completed_at
              FROM ingestion_log
             WHERE source_id IN ({placeholders})
            """
        )
        with sqlite3.connect(self.sqlite_path) as connection:
            connection.row_factory = sqlite3.Row
            rows = connection.execute(query, tuple(valid_sources)).fetchall()

        lookup: Dict[Tuple[str, str | None], Dict[str, object]] = {}
        latest: Dict[str, Dict[str, object]] = {}
        for row in rows:
            info = {
                "run_id": row["run_id"],
                "source_id": row["source_id"],
                "url": row["url"],
                "local_path": row["local_path"],
                "checksum": row["checksum"],
                "completed_at": row["completed_at"],
            }
            lookup[(row["source_id"], row["run_id"])] = info
            latest_entry = latest.get(row["source_id"])
            if (
                latest_entry is None
                or str(row["completed_at"]) > str(latest_entry.get("completed_at"))
            ):
                latest[row["source_id"]] = info
        for source_id, info in latest.items():
            lookup.setdefault((source_id, None), info)
        return lookup


__all__ = ["ExportGenerator", "ExportSummary"]
