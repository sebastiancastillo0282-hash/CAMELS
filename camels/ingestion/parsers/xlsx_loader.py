"""XLSX parser for ingestion."""
from __future__ import annotations

from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook

from .base import ParsedDataset


def parse_xlsx(path: Path, *, worksheet: str | None = None) -> ParsedDataset:
    workbook = load_workbook(path, data_only=True, read_only=True)
    if worksheet:
        if worksheet not in workbook.sheetnames:
            raise ValueError(f"Worksheet '{worksheet}' not found in {path.name}")
        sheet = workbook[worksheet]
    else:
        sheet = workbook.active
    rows = list(sheet.rows)
    if not rows:
        return ParsedDataset(records=[], metadata={"columns": [], "worksheet": sheet.title})
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in rows[0]]
    records: List[Dict[str, object]] = []
    for row in rows[1:]:
        values = [cell.value for cell in row]
        record = {headers[index] if index < len(headers) else f"column_{index}": value for index, value in enumerate(values)}
        records.append(record)
    metadata = {
        "columns": headers,
        "worksheet": sheet.title,
    }
    return ParsedDataset(records=records, metadata=metadata)
